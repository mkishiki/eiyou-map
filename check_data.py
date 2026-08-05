#!/usr/bin/env python3
"""データ整合チェッカー（data-map v1「レベル2」）

食材・栄養素を足したあとに実行して、取りこぼしを機械的に見つける。
gen_pct.py / build.py の前に走らせる想定。判定はするが、修正はしない。

使い方:  python3 check_data.py
入力:    amounts.csv, tabeawase-data.js, 成分表Excel（下記 XLSX を実ファイル名に合わせる）
出力:    3種の検査結果。異常があれば非ゼロ終了。

検査:
  1. 閾値もれ … 成分表で一食10%以上なのに FOODS に無い紐付き（＝手書きの入れ忘れ）
  2. 整合    … amounts.csv と FOODS の食材×栄養素が一致するか（PKズレ・幽霊リンク）
  3. 孤立    … data-rule 4 の救済。どの栄養素とも紐付かない食材がいないか

data-rule の閾値・救済ルールが唯一の根拠。ここを変えたら data-rule も直す。
"""
import csv, re, sys, pathlib

HERE = pathlib.Path(__file__).parent
CSV  = HERE / "amounts.csv"
JS   = HERE / "tabeawase-data.js"
XLSX = HERE / "20260327mxt_kagseimext000029402_02.xlsx"  # 成分表（増補2023・正誤反映）

THRESHOLD = 10.0  # data-rule 4：一食が一日目安の10%以上で紐付ける

MAP = {
    'b1':'THIA','b12':'VITB12','b2':'RIBF','b6':'VITB6A','ca':'CA','cu':'CU',
    'fat':'FAT-','fe':'FE','fiber':'FIB-','folate':'FOL','k':'K','mg':'MG',
    'na':'NACL_EQ','niacin':'NE','p':'P','panto':'PANTAC','protein':'PROT-',
    'se':'SE','sugar':'CHOAVLM','va':'VITA_RAE','vc':'VITC','vd':'VITD',
    've':'TOCPHA','vk':'VITK','zn':'ZN','alc':'ALC',
}
SKIP_MISSING = {'sugar', 'alc'}  # 主成分・別基準。閾値もれ検査から除外


def load_csv():
    foods, daily, name, linked = {}, {}, {}, set()
    with open(CSV, encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            food, nid, code = row['食材'], row['栄養素id'], row['食品番号'].strip()
            foods.setdefault(food, {'g': float(row['一食の目安量g']), 'code': code})
            daily.setdefault(nid, float(row['一日の目安量']))
            name.setdefault(nid, row['栄養素'])
            linked.add((nid, food))
    return foods, daily, name, linked


def load_foods_links():
    js = JS.read_text(encoding='utf-8')
    m = re.search(r'const FOODS=\[(.*?)\n\];', js, re.S)
    if not m:
        sys.exit('tabeawase-data.js に const FOODS=[...] が見つからない')
    return set(re.findall(r'\["([a-z0-9]+)","([^"]+)",\d', m.group(1)))


def load_excel():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    hdr = [str(v).strip() if v else '' for v in
           list(wb[wb.sheetnames[1]].iter_rows(min_row=12, max_row=12, values_only=True))[0]]
    col = {h: i for i, h in enumerate(hdr) if h}
    def num(v):
        if v is None: return None
        s = str(v).replace('Tr', '0').replace('(', '').replace(')', '').strip()
        if s in ('', '-'): return None
        try: return float(s)
        except: return None
    db = {}
    for sh in wb.sheetnames[1:]:
        for row in wb[sh].iter_rows(min_row=12, values_only=True):
            code = row[1]
            if code is None: continue
            code = str(code).strip().zfill(5)
            if code.isdigit():
                db[code] = {k: num(row[i]) for k, i in col.items()}
    return db


def main():
    foods, daily, name, linked_csv = load_csv()
    linked_foods = load_foods_links()
    ok = True

    print('■ 検査2: 整合（csv <-> FOODS）')
    print(f'  amounts.csv 紐付き: {len(linked_csv)}')
    print(f'  FOODS 紐付き:       {len(linked_foods)}')
    only_csv = linked_csv - linked_foods
    only_foods = linked_foods - linked_csv
    if only_csv:
        ok = False
        print(f'  x csv にあるが FOODS に無い: {len(only_csv)}')
        for k in sorted(only_csv)[:20]: print(f'      {k[0]}|{k[1]}')
    if only_foods:
        ok = False
        print(f'  x FOODS にあるが csv に無い（幽霊リンク）: {len(only_foods)}')
        for k in sorted(only_foods)[:20]: print(f'      {k[0]}|{k[1]}')
    if not only_csv and not only_foods:
        print('  OK 一致')

    print('\n■ 検査1: 閾値もれ（一食10%以上なのに FOODS に無い）')
    db = load_excel()
    missing = []
    for food, info in foods.items():
        code = info['code'].zfill(5) if info['code'] else ''
        if not code or code not in db: continue  # 食品番号なし（総称等）は対象外
        g, rec = info['g'], db[code]
        for nid, d in daily.items():
            if nid in SKIP_MISSING: continue
            if (nid, food) in linked_foods: continue
            sid = MAP.get(nid)
            if not sid or sid not in rec: continue
            val = rec[sid]
            if val is None: continue
            pct = val * g / 100 / d * 100
            if pct >= THRESHOLD:
                missing.append((food, name[nid], round(pct, 1)))
    if missing:
        ok = False
        missing.sort(key=lambda x: -x[2])
        print(f'  x 取りこぼし {len(missing)} 件:')
        for food, nm, pct in missing:
            print(f'      {food} x {nm}  {pct}%')
    else:
        print('  OK なし')

    print('\n■ 検査3: 孤立（どの栄養素とも紐付かない食材）')
    linked_names = set(f for _, f in linked_foods)
    orphan = [f for f in foods if f not in linked_names]
    if orphan:
        ok = False
        print(f'  x 孤立 {len(orphan)} 件: {orphan}')
        print('    （data-rule 4 の救済：最低1本は紐付ける）')
    else:
        print('  OK なし')

    print('\n' + ('=== すべて OK ===' if ok else '=== 要修正あり ==='))
    sys.exit(0 if ok else 1)


if __name__ == '__main__':
    main()
